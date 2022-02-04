import time

from screen import layout, login_layout
import PySimpleGUI as sg
from worker import AutoWorker
from openpyxl import load_workbook
from datetime import datetime
# import json


def get_data(path):
    wb = load_workbook(path)
    ws = wb.active
    col_nr = 0
    length = 0
    for i in range(1, 100):
        if ws.cell(row=1, column=i).value == "Schaufenster-link-href":
            col_nr = i
            break

    for i in range(1, 1000000):
        if ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=1).value == "":
            length = i
            break

    data = []

    for i in range(2, length):
        url = ws.cell(row=i, column=col_nr).value
        if url != "" and url is not None:
            data.append(url)
    wb.close()
    return data, col_nr, path, length


def save_data(path, n_col, data, length):
    wb = load_workbook(path)
    ws = wb.active
    ws.insert_cols(n_col + 1)
    counter = 0
    ws.cell(row=1, column=n_col + 1).value = "Total Revenues"
    for i in range(2, length):
        if ws.cell(row=i, column=n_col).value is not None and ws.cell(row=i, column=n_col).value != "":
            ws.cell(row=i, column=n_col + 1).value = data[counter]
            counter += 1

    wb.save(path)
    wb.close()


def action():
    window = sg.Window("Helium Automation", layout)
    event, values = window.read()
    while True:
        if event == sg.WINDOW_CLOSED:
            window.close()
            break
        if event == "Cancel":
            window.close()
            break
        if event == "Test":
            test_res = AutoWorker().test_confidence("https://www.amazon.de/-/en/s?ie=UTF8&marketplaceID=A1PA6795UKMFR9&me=A2Y45U1SJYRN75")
            print(test_res)
            time.sleep(2)
            continue

        if event == "Start":
            worker = AutoWorker()
            if not worker.test_if_logged_in():
                win2 = sg.Window("Login", login_layout)
                ev2, val2 = win2.read()
                while True:
                    if ev2 == sg.WINDOW_CLOSED:
                        exit()
                    if ev2 == "Cancel":
                        exit()
                    if ev2 == "Login":
                        user = val2["user"]
                        password = val2["pass"]
                        if worker.login(user, password):
                            win2.close()
                            break
                        else:
                            sg.PopupError("Sorry etwas hat nicht funktioniert")
                            exit()
            urls, row_nr, path, total_len = get_data(values["file"])
            start_time = datetime.now()
            print(urls)
            total_revs = []
            worker.activate()
            for li in urls:
                res = worker.get_total_rev(li)
                if res == "0":
                    res = worker.get_total_rev(li, long=True)
                total_revs.append(res)
            worker.deactivate()
            print(total_revs)
            save_data(path, row_nr, total_revs, total_len)
            total_time = datetime.now() - start_time
            sg.PopupOK(
                f"Fertig. Benötigte Zeit: {total_time.seconds // 3600} Stunden und {(total_time.seconds // 60) % 60} Minuten bei {len(urls)} Links")
            exit()
