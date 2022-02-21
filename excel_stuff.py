import openpyxl


def get_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    col_nr = 0
    length = 0
    for i in range(1, 100):
        if ws.cell(row=1, column=i).value == "Schaufenster-link-href-2":
            col_nr = i
            break

    for i in range(1, 1000000):
        if ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=1).value == "":
            length = i
            break

    data = []

    for i in range(2, length):
        url = ws.cell(row=i, column=col_nr).value
        if url != "" and url is not None:
            data.append(url)
    wb.close()
    return data, col_nr, path, length


def save_data(path, n_col, data, length):
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    ws.insert_cols(n_col + 1)
    counter = 0
    ws.cell(row=1, column=n_col + 1).value = "Total Revenues"
    for i in range(2, length):
        if ws.cell(row=i, column=n_col).value is not None and ws.cell(row=i, column=n_col).value != "":
            ws.cell(row=i, column=n_col + 1).value = data[counter]
            counter += 1

    wb.save(path)
    wb.close()


def check_for_fails(data: list[str]) -> list[int]:
    # get indexes of every "0" el (not found)
    null_indexes = [i for i in range(len(data)) if data[i] == "0"]
    elements_to_del = []
    counter = 2
    for null_ind in null_indexes[2:-2]:
        if null_indexes[counter - 1] == (null_ind - 1) and null_indexes[counter - 2] == (null_ind - 2):
            counter += 1
            continue
        elif null_indexes[counter + 1] == (null_ind + 1) and null_indexes[counter + 2] == (null_ind + 2):
            counter += 1
            continue
        else:
            counter += 1
            elements_to_del.append(null_ind)

    for el in elements_to_del:
        null_indexes.remove(el)

    return null_indexes[2:]
